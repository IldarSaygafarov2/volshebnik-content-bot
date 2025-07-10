from pprint import pprint

import openpyxl
from settings import HEADERS

PHOTOS_SITE_DOMAIN = 'robins.ru'


def get_data_from_excel_file(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    my_list = list()
    for idx, value in enumerate(
        ws.iter_rows(
            min_row=1, max_row=ws.max_row + 1, min_col=1, max_col=13, values_only=True
        )
    ):
        is_all_none = all(list(map(lambda x: x is None, value[1:])))
        if not is_all_none:
            if idx == 0:
                continue
            print(dict(zip(HEADERS, value)))
            my_list.append(dict(zip(HEADERS, value)))

    return my_list



