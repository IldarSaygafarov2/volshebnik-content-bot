import openpyxl


wb = openpyxl.load_workbook("data.xlsx")

ws = wb.active


my_list = list()

for value in ws.iter_rows(
    min_row=1, max_row=ws.max_row + 1, min_col=1, max_col=13, values_only=True
):
    print(value)
    print(list(filter(lambda x: x is not None, value)))


# for ele1, ele2, ele3, ele4, ele5, ele6 in my_list:
#     print(
#         "{:<8}{:<35}{:<10}s{:<10}{:<15}{:<15}".format(
#             ele1, ele2, ele3, ele4, ele5, ele6
#         )
#     )
