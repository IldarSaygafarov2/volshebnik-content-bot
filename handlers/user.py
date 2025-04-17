from data.loader import bot
from telebot.types import Message
import openpyxl


@bot.message_handler(commands=["start"])
def handle_command_start(message: Message):
    chat_id = message.from_user.id

    bot.send_message(chat_id, "Отправьте эксель файл с обновленными данными")


@bot.message_handler(content_types=["document"])
def recieve_data_from_excel_file(message: Message):
    print(message.document)
    document_id = message.document.file_id

    _file = bot.get_file(document_id)
    file = bot.download_file(_file.file_path)

    with open(f"data.xlsx", mode="wb") as f:
        f.write(file)

    wb = openpyxl.load_workbook(f"data.xlsx")
    ws = wb.active
    # values = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    # print(values)
    print(ws.rows)
    print(ws.columns)
